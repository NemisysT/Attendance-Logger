import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import calendar
import tkinter as tk
from tkinter import filedialog, simpledialog
from tkinter import ttk
import logging
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# Employee data dictionary (name as key, ID as value)
employee_data = {
    #Enter the Data in the following Format:
    "John": 1,
    "Eve":2,
}
# Logging configuration
logging.basicConfig(filename="attendance_log.txt", level=logging.INFO, format="%(asctime)s - %(message)s")

def select_holidays(month_year):
    """Ask the user to select holidays for the given month."""
    root = tk.Tk()
    root.withdraw()
    days_in_month = calendar.monthrange(int(month_year.split("-")[1]), list(calendar.month_name).index(month_year.split("-")[0]))[1]
    holidays = simpledialog.askstring(
        "Input Holidays",
        f"Enter holiday dates for the month, comma-separated, e.g., 1,15,28:"
    )
    if holidays:
        return {int(day.strip()) for day in holidays.split(",") if day.strip().isdigit() and 1 <= int(day.strip()) <= days_in_month}
    return set()

def calculate_work_hours(punches):
    """
    Calculate total work hours in HH:MM format based on punch times.
    Handles multiple punch pairs (e.g., in-out, in-out).
    """
    if len(punches) < 2 or len(punches) % 2 != 0:
        return "N.A."  # Not enough punches or uneven number of punches

    total_work_seconds = 0

    # Iterate over punches in pairs (start, end)
    for i in range(0, len(punches), 2):
        work_interval = punches[i + 1] - punches[i]
        if work_interval < timedelta(0):
            return "N.A"  # Ensures no negative intervals
        total_work_seconds += work_interval.total_seconds()

    # Convert total seconds into hours and minutes
    total_minutes = total_work_seconds // 60
    hours = int(total_minutes // 60)
    minutes = int(total_minutes % 60)
    return f"{hours:02}:{minutes:02}"  # Format as HH:MM

    

def process_attendance(file_path):
    """Process the attendance file and generate the Excel report."""
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Extract the month and year from the first valid date
    for line in lines:
        parts = line.strip().split()
        if len(parts) < 2:
            continue
        try:
            timestamp = datetime.strptime(parts[1], "%Y-%m-%d")
            month_year = timestamp.strftime("%B-%Y")
            break
        except ValueError:
            continue

    holidays = select_holidays(month_year)
    days_in_month = calendar.monthrange(timestamp.year, timestamp.month)[1]
    attendance_data = {emp_id: {} for emp_id in employee_data.values()}

    # Parse the file for employee attendance
    for line in lines:
        parts = line.strip().split()
        if len(parts) < 3:  # Ensure we have at least employee ID, date, and time
            continue

        try:
            # Parse using the first three parts (employee ID, date, time)
            emp_id = int(parts[0])
            timestamp = datetime.strptime(f"{parts[1]} {parts[2]}", "%Y-%m-%d %H:%M:%S")
            
            if emp_id in attendance_data:
                if timestamp.day not in attendance_data[emp_id]:
                    attendance_data[emp_id][timestamp.day] = []
                attendance_data[emp_id][timestamp.day].append(timestamp)
        except ValueError:
            logging.warning(f"Skipping malformed line: {line}")
            continue

    # Determine Sundays for the given month
    first_day = datetime(timestamp.year, timestamp.month, 1)
    sundays = {d for d in range(1, days_in_month + 1) if (first_day + timedelta(days=d - 1)).weekday() == 6}

    # Create the Excel report
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    day_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick_border = Border(
        left=Side(style='thick', color="000000"),
        right=Side(style='thick', color="000000"),
        top=Side(style='thick', color="000000"),
        bottom=Side(style='thick', color="000000")
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define colors for "P" (present) and "AB" (absent)
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green
    absent_fill = PatternFill(start_color="F9C1C1", end_color="F9C1C1", fill_type="solid")  # Light Red
    late_punch_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light Yellow
    holiday_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Light Gray

    # Header row with month-year
    ws['A1'] = f"Month: {month_year}"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    ws['A1'].fill = header_fill
    ws['A1'].border = thick_border

    # Adjust column widths for better alignment
    ws.column_dimensions['A'].width = 30  # Name column
    
    # Add row for days of the week (MON, TUES, etc.)
    for day in range(1, days_in_month + 1):
        weekday = (first_day + timedelta(days=day - 1)).strftime("%a").upper()
        cell = ws.cell(row=2, column=day + 1, value=weekday)
        cell.alignment = center_alignment
        cell.fill = day_header_fill
        cell.font = Font(bold=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(day + 1)].width = 5  # Set narrow width for day columns

    # Employee names in the third row
    sorted_names = sorted(employee_data.keys())
    for idx, name in enumerate(sorted_names, start=3):
        cell = ws[f"A{idx}"]
        cell.value = name
        cell.alignment = center_alignment
        cell.border = thin_border

    # Days of the month as columns, showing the full date (e.g., 1/11/24)
    for day in range(1, days_in_month + 1):
        full_date = datetime(timestamp.year, timestamp.month, day).strftime(f"{day}/%m/%y")
        cell = ws.cell(row=1, column=day + 1, value=full_date)
        cell.alignment = center_alignment
        cell.fill = day_header_fill
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Total column header
    total_col = days_in_month + 2
    cell = ws.cell(row=1, column=total_col, value="Total")
    cell.font = Font(bold=True)
    cell.alignment = center_alignment
    cell.fill = header_fill
    cell.border = thick_border

    # Fill attendance
    for idx, name in enumerate(sorted_names, start=3):
        emp_id = employee_data[name]
        total_present = 0
        for day in range(1, days_in_month + 1):
            col = day + 1
            cell = ws.cell(row=idx, column=col)
            
            # Check if the day is a holiday
            if day in holidays:
                cell.value = "H"  # Holiday marker
                cell.fill = holiday_fill
            elif day in sundays:
                # If it's a Sunday, mark "P" for attendance if punched in
                if day in attendance_data[emp_id]:
                    cell.value = "P"
                    cell.fill = present_fill
                    total_present += 1
                else:
                    cell.value = ""  # Leave blank if no punch
            elif day in attendance_data[emp_id]:
                # Regular working day
                cell.value = "P"
                cell.fill = present_fill
                total_present += 1
            else:
                cell.value = "AB"
                cell.fill = absent_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Fill total present days for this employee in the total column
        total_cell = ws.cell(row=idx, column=total_col, value=total_present)
        total_cell.alignment = center_alignment
        total_cell.border = thin_border

    # Detailed Punch Report for each employee
    report_start_row = len(employee_data) + 5

    # Individual Report for Each Employee
    for idx, (emp_name, emp_id) in enumerate(employee_data.items(), 1):
        # Check if employee has any punches
        if not attendance_data[emp_id]:
            # No punches at all - create an AB entry
            report_row = report_start_row + (idx - 1) * (days_in_month + 4)
            
            # Enhanced Header
            ws.cell(row=report_row, column=1, value=f"Attendance Report for {emp_name}")
            ws.merge_cells(start_row=report_row, start_column=1, end_row=report_row, end_column=7)
            header_cell = ws[report_row][1]
            header_cell.font = Font(bold=True, size=14, color="FFFFFF")
            header_cell.alignment = center_alignment
            header_cell.fill = header_fill
            header_cell.border = thick_border

            # Table Headers with improved styling
            headers = ["Day", "Punch 1", "Punch 2", "Punch 3", "Punch 4", "Hours", "Status"]
            for col, header in enumerate(headers, 1):
                header_cell = ws.cell(row=report_row + 1, column=col, value=header)
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.alignment = center_alignment
                header_cell.fill = subheader_fill
                header_cell.border = thick_border

            # Fill AB for all days with improved styling
            for day in range(1, days_in_month + 1):
                row = report_row + 1 + day
                weekday = (first_day + timedelta(days=day - 1)).weekday()
                
                # Day column
                day_cell = ws.cell(row=row, column=1, value=day)
                day_cell.alignment = center_alignment
                day_cell.border = thin_border
                
                # Punch columns
                for col in range(2, 6):
                    punch_cell = ws.cell(row=row, column=col)
                    punch_cell.border = thin_border
                
                # Status column
                if day in holidays or weekday == 6:  # Sunday or Holiday
                    status_cell = ws.cell(row=row, column=7, value="H")
                    status_cell.fill = holiday_fill
                else:
                    status_cell = ws.cell(row=row, column=7, value="AB")
                    status_cell.fill = absent_fill
                status_cell.alignment = center_alignment
                status_cell.border = thin_border

            continue

        # Create Report Header
        report_row = report_start_row + (idx - 1) * (days_in_month + 4)
        
        # Enhanced Header
        ws.cell(row=report_row, column=1, value=f"Attendance Report for {emp_name}")
        ws.merge_cells(start_row=report_row, start_column=1, end_row=report_row, end_column=7)
        header_cell = ws[report_row][1]
        header_cell.font = Font(bold=True, size=14, color="FFFFFF")
        header_cell.alignment = center_alignment
        header_cell.fill = header_fill
        header_cell.border = thick_border

        # Table Headers with improved styling
        headers = ["Day", "Punch 1", "Punch 2", "Punch 3", "Punch 4", "Hours", "Status"]
        for col, header in enumerate(headers, 1):
            header_cell = ws.cell(row=report_row + 1, column=col, value=header)
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.alignment = center_alignment
            header_cell.fill = subheader_fill
            header_cell.border = thick_border

        # Fill Punch Details for Each Day
        for day in range(1, days_in_month + 1):
            row = report_row + 1 + day
            weekday = (first_day + timedelta(days=day - 1)).weekday()
            
            # Day column
            day_cell = ws.cell(row=row, column=1, value=day)
            day_cell.alignment = center_alignment
            day_cell.border = thin_border
            
            if day in attendance_data[emp_id]:
                # Sort punches for the day
                day_punches = sorted(attendance_data[emp_id][day])
                
                # Fill punch times
                for punch_idx, punch in enumerate(day_punches, start=2):
                    punch_time = punch.strftime("%H:%M:%S")
                    punch_cell = ws.cell(row=row, column=punch_idx, value=punch_time)
                    punch_cell.alignment = center_alignment
                    punch_cell.border = thin_border
                    
                    # Late punch highlighting
                    if punch_idx == 2 and punch.time() > datetime.strptime("09:05:00", "%H:%M:%S").time():
                        punch_cell.fill = late_punch_fill
                    elif punch_idx == 4 and punch.time() > datetime.strptime("13:35:00", "%H:%M:%S").time():
                        punch_cell.fill = late_punch_fill

                # Fill remaining punch columns if fewer than 4 punches
                for col in range(len(day_punches) + 2, 6):
                    empty_cell = ws.cell(row=row, column=col)
                    empty_cell.border = thin_border

                # Calculate work hours
                work_hours = calculate_work_hours(day_punches)
                hours_cell = ws.cell(row=row, column=6, value=work_hours)
                hours_cell.alignment = center_alignment
                hours_cell.border = thin_border

                # Status cell
                status_cell = ws.cell(row=row, column=7, value="P")
                status_cell.fill = present_fill
                status_cell.alignment = center_alignment
                status_cell.border = thin_border
            else:
                # No punch - mark based on day type
                for col in range(2, 6):
                    empty_cell = ws.cell(row=row, column=col)
                    empty_cell.border = thin_border
                
                # Check if day is a holiday or Sunday
                if day in holidays or weekday == 6:
                    status_cell = ws.cell(row=row, column=7, value="H")
                    status_cell.fill = holiday_fill
                else:
                    status_cell = ws.cell(row=row, column=7, value="AB")
                    status_cell.fill = absent_fill
                
                status_cell.alignment = center_alignment
                status_cell.border = thin_border

        # Adjust column widths for better readability
        column_widths = [5, 10, 10, 10, 10, 8, 8]  # Adjusted to match 7 columns
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

    # Save the workbook
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if output_file:
        wb.save(output_file)
        logging.info(f"Attendance report saved to {output_file}")
        print(f"Attendance report saved to {output_file}")
    else:
        print("Report generation canceled.")

def select_file():
    """File selection dialog with an improved GUI."""
    def open_file():
        file_path = filedialog.askopenfilename(
            title="Select Attendance File",
            filetypes=[("Text Files", "*.dat")]
        )
        if file_path:
            status_label.config(text=f"Selected: {file_path}")
            process_attendance(file_path)
        else:
            status_label.config(text="No file selected.")

    # Main window
    root = tk.Tk()
    root.title("Met Air Attendance Report")
    root.geometry("400x200")
    root.resizable(False, False)

    # Style
    style = ttk.Style()
    style.theme_use("clam")

    # Header
    header_label = ttk.Label(
        root,
        text="Generate Attendance Report",
        font=("Arial", 16, "bold"),
        anchor="center"
    )
    header_label.pack(pady=10)

    # Buttons and labels
    button_frame = ttk.Frame(root)
    button_frame.pack(pady=10)

    open_button = ttk.Button(button_frame, text="Select Attendance File", command=open_file)
    open_button.grid(row=0, column=0, padx=5, pady=5)

    quit_button = ttk.Button(button_frame, text="Exit", command=root.destroy)
    quit_button.grid(row=0, column=1, padx=5, pady=5)

    # Status Label
    status_label = ttk.Label(root, text="No file selected.", anchor="center")
    status_label.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    select_file()